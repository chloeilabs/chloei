#!/usr/bin/env python3
import argparse
import json
import os
import re
import sys
import zipfile
from pathlib import Path

try:
    from defusedxml import ElementTree as SafeElementTree
except ImportError:
    from xml.etree import ElementTree as UnsafeElementTree

    class SafeElementTree:
        @staticmethod
        def fromstring(value):
            if re.search(rb"<!\s*(?:DOCTYPE|ENTITY)\b", value, re.IGNORECASE):
                raise ValueError("Unsafe XML document type or entity declaration.")
            return UnsafeElementTree.fromstring(value)


TEXT_EXTENSIONS = {
    ".csv",
    ".htm",
    ".html",
    ".ipynb",
    ".json",
    ".md",
    ".overpassql",
    ".py",
    ".txt",
    ".yaml",
    ".yml",
}


def clean_text(value):
    value = re.sub(r"\r\n?", "\n", str(value))
    value = re.sub(r"[ \t]+", " ", value)
    value = re.sub(r"\n{4,}", "\n\n\n", value)
    return value.strip()


def cap_text(value, max_chars):
    value = clean_text(value)
    if len(value) <= max_chars:
        return value, False
    return (
        value[:max_chars]
        + f"\n\n[TRUNCATED: kept {max_chars} of {len(value)} characters]",
        True,
    )


def local_name(tag):
    return tag.rsplit("}", 1)[-1]


def xml_text(node, text_tags):
    parts = []
    for child in node.iter():
        name = local_name(child.tag)
        if name in text_tags and child.text:
            parts.append(child.text)
        elif name == "tab":
            parts.append("\t")
        elif name in {"br", "cr"}:
            parts.append("\n")
    return clean_text("".join(parts))


def extract_text_file(path, max_chars):
    if path.suffix.lower() == ".ipynb":
        data = json.loads(path.read_text("utf-8", errors="replace"))
        lines = []
        for index, cell in enumerate(data.get("cells", []), start=1):
            cell_type = cell.get("cell_type", "unknown")
            source = "".join(cell.get("source", []))
            lines.append(f"## Cell {index}: {cell_type}\n{source}")
        return cap_text("\n\n".join(lines), max_chars)

    return cap_text(path.read_text("utf-8", errors="replace"), max_chars)


def extract_docx(path, max_chars):
    lines = []
    with zipfile.ZipFile(path) as archive:
        document = SafeElementTree.fromstring(archive.read("word/document.xml"))
    body = next((child for child in document if local_name(child.tag) == "body"), None)
    if body is None:
        return "", False

    table_index = 0
    for child in body:
        name = local_name(child.tag)
        if name == "p":
            text = xml_text(child, {"t"})
            if text:
                lines.append(text)
        elif name == "tbl":
            table_index += 1
            lines.append(f"\n[TABLE {table_index}]")
            for row in [node for node in child if local_name(node.tag) == "tr"]:
                cells = []
                for cell in [node for node in row if local_name(node.tag) == "tc"]:
                    cells.append(xml_text(cell, {"t"}).replace("\n", " "))
                if any(cells):
                    lines.append("\t".join(cells))

    return cap_text("\n".join(lines), max_chars)


def slide_sort_key(name):
    match = re.search(r"slide(\d+)\.xml$", name)
    return int(match.group(1)) if match else 0


def extract_pptx(path, max_chars):
    sections = []
    with zipfile.ZipFile(path) as archive:
        names = sorted(
            [name for name in archive.namelist() if re.search(r"ppt/slides/slide\d+\.xml$", name)],
            key=slide_sort_key,
        )
        for index, name in enumerate(names, start=1):
            root = SafeElementTree.fromstring(archive.read(name))
            text = xml_text(root, {"t"})
            if text:
                sections.append(f"## Slide {index}\n{text}")
            else:
                sections.append(f"## Slide {index}\n[No extracted text]")
    return cap_text("\n\n".join(sections), max_chars)


def extract_pdf(path, max_chars):
    try:
        from pypdf import PdfReader
    except Exception as exc:
        return f"[PDF extraction unavailable: pypdf import failed: {exc}]", False

    reader = PdfReader(str(path))
    sections = []
    for index, page in enumerate(reader.pages, start=1):
        try:
            text = page.extract_text() or ""
        except Exception as exc:
            text = f"[Page extraction failed: {exc}]"
        if text.strip():
            sections.append(f"## Page {index}\n{text}")
        if sum(len(section) for section in sections) >= max_chars:
            break
    return cap_text(f"Pages: {len(reader.pages)}\n\n" + "\n\n".join(sections), max_chars)


def extract_zip_listing(path, max_chars):
    with zipfile.ZipFile(path) as archive:
        rows = []
        for info in archive.infolist():
            rows.append(f"{info.filename}\t{info.file_size} bytes")
    return cap_text("ZIP contents:\n" + "\n".join(rows), max_chars)


def cell_to_text(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return value.replace("\n", " ").strip()
    return str(value)


def extract_xlsx(path, max_chars, xlsx_max_rows, xlsx_max_cols, xlsx_formula_limit):
    try:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
    except Exception as exc:
        return f"[XLSX extraction unavailable: openpyxl import failed: {exc}]", False

    lines = []
    formula_limit = xlsx_formula_limit
    table_row_limit = xlsx_max_rows
    table_col_limit = xlsx_max_cols

    wb_formulas = load_workbook(path, read_only=True, data_only=False, keep_links=False)
    wb_values = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    lines.append(f"Workbook sheets: {', '.join(wb_formulas.sheetnames)}")

    for sheet_name in wb_formulas.sheetnames:
        ws = wb_formulas[sheet_name]
        ws_values = wb_values[sheet_name]
        try:
            ws.calculate_dimension(force=True)
            ws_values.calculate_dimension(force=True)
        except TypeError:
            ws.calculate_dimension()
            ws_values.calculate_dimension()
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        shown_rows = min(max_row, table_row_limit)
        shown_cols = min(max_col, table_col_limit)
        lines.append(
            f"\n## Sheet: {sheet_name}\nDimension: {max_row} rows x {max_col} columns; "
            f"preview: first {shown_rows} rows x {shown_cols} columns"
        )

        if shown_rows > 0 and shown_cols > 0:
            header = ["row"] + [get_column_letter(col) for col in range(1, shown_cols + 1)]
            lines.append("\t".join(header))
            formula_count = 0
            formulas = []
            value_rows = ws_values.iter_rows(
                min_row=1,
                max_row=shown_rows,
                min_col=1,
                max_col=shown_cols,
            )
            formula_rows = ws.iter_rows(
                min_row=1,
                max_row=shown_rows,
                min_col=1,
                max_col=shown_cols,
            )
            for row_number, (formula_row, value_row) in enumerate(
                zip(formula_rows, value_rows), start=1
            ):
                values = []
                for formula_cell, value_cell in zip(formula_row, value_row):
                    value = formula_cell.value
                    cached_value = value_cell.value
                    if isinstance(value, str) and value.startswith("="):
                        formula_count += 1
                        if len(formulas) < formula_limit:
                            formulas.append(
                                f"{formula_cell.coordinate}: {value} -> cached {cell_to_text(cached_value)}"
                            )
                        values.append(f"{value} [cached: {cell_to_text(cached_value)}]")
                    else:
                        values.append(cell_to_text(value))
                if any(values):
                    lines.append("\t".join([str(row_number), *values]))

            if formulas:
                lines.append(f"\nFormulas observed in preview: {formula_count}")
                lines.extend(formulas)
                if formula_count > len(formulas):
                    lines.append(
                        f"[Formula list truncated: kept {len(formulas)} of {formula_count}]"
                    )

        if len("\n".join(lines)) >= max_chars:
            break

    return cap_text("\n".join(lines), max_chars)


def extract(path, max_chars, xlsx_max_rows, xlsx_max_cols, xlsx_formula_limit):
    ext = path.suffix.lower()
    if ext == ".xlsx":
        return "xlsx", extract_xlsx(
            path,
            max_chars,
            xlsx_max_rows,
            xlsx_max_cols,
            xlsx_formula_limit,
        )
    if ext == ".docx":
        return "docx", extract_docx(path, max_chars)
    if ext == ".pptx":
        return "pptx", extract_pptx(path, max_chars)
    if ext == ".pdf":
        return "pdf", extract_pdf(path, max_chars)
    if ext == ".zip":
        return "zip", extract_zip_listing(path, max_chars)
    if ext in TEXT_EXTENSIONS:
        return "text", extract_text_file(path, max_chars)
    if ext in {".png", ".jpg", ".jpeg"}:
        return "image", (f"[Image file: {path.name}; OCR is not implemented]", False)
    return "unsupported", (f"[Unsupported file extension: {ext or '(none)'}]", False)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--url", default="")
    parser.add_argument("--max-chars", type=int, default=120000)
    parser.add_argument("--xlsx-max-rows", type=int, default=250)
    parser.add_argument("--xlsx-max-cols", type=int, default=50)
    parser.add_argument("--xlsx-formula-limit", type=int, default=400)
    args = parser.parse_args()

    path = Path(args.input)
    record = {
        "fileName": path.name,
        "extension": path.suffix.lower(),
        "url": args.url,
        "sizeBytes": os.path.getsize(path),
        "status": "ok",
        "kind": "unknown",
        "truncated": False,
        "text": "",
        "warnings": [],
    }

    try:
        kind, (text, truncated) = extract(
            path,
            args.max_chars,
            max(1, args.xlsx_max_rows),
            max(1, args.xlsx_max_cols),
            max(0, args.xlsx_formula_limit),
        )
        record["kind"] = kind
        record["text"] = text
        record["truncated"] = truncated
        if kind in {"image", "unsupported"} or "extraction unavailable" in text.lower():
            record["status"] = "limited"
    except Exception as exc:
        record["status"] = "error"
        record["kind"] = path.suffix.lower().lstrip(".") or "unknown"
        record["warnings"].append(f"{type(exc).__name__}: {exc}")
        record["text"] = f"[Extraction failed: {type(exc).__name__}: {exc}]"

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(record, indent=2) + "\n", encoding="utf-8")
    print(json.dumps({"output": str(output_path), "status": record["status"]}))


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"extract-gdpval-file failed: {exc}", file=sys.stderr)
        raise
